import os
from openpyxl import load_workbook

# 加载Excel文件
workbook = load_workbook('./dates.xlsx')
sheet = workbook.active

# 获取当前文件夹路径
current_folder = os.getcwd()

# 图片文件所在的文件夹路径
image_folder = 'C:/Users/Administrator/Desktop/3/images'

# 遍历图片文件夹中的所有图片文件
for filename in os.listdir(image_folder):
    # 检查文件是否为图片文件
    if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
        # 读取当前行的数据
        row_index = int(os.path.splitext(filename)[0].split('_')[1])
        data_a1 = sheet.cell(row=row_index, column=1).value
        data_b2 = sheet.cell(row=row_index, column=2).value
        
        # 检查数据是否为空
        if data_a1 is None or data_b2 is None:
            print(f'第 {row_index} 行的数据为空。无法更改文件名。')
            continue
        
        # 构建完整的旧文件路径和新文件路径
        old_path = os.path.join(image_folder, filename)
        new_name = str(data_a1) + str(data_b2) + os.path.splitext(filename)[1]
        new_path = os.path.join(image_folder, new_name)

        # 重命名文件
        os.rename(old_path, new_path)
        print(f'图片文件 {filename} 已重命名为 {new_name}')

print('图片文件重命名完成。')
