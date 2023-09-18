from openpyxl import Workbook
from datetime import datetime, timedelta
from random import randint

# 创建一个工作簿
workbook = Workbook()
# 选择默认的活动工作表
sheet = workbook.active

# 设置日期格式
date_format = '%Y/%m/%d'
time_format = '%H:%M:%S'

# 设置初始日期和时间
start_date = datetime(2022, 5, 1)

# 设置结束日期和时间
end_date = datetime(2023, 5, 31)

# 写入日期和时间数据
current_date = start_date.date()
row_index = 1
while current_date <= end_date.date():
    # 在早上8点到晚上8点之间随机生成三个时分秒数据
    for i in range(1):
        random_hour = randint(8, 20)
        random_minute = randint(0, 59)
        random_second = randint(0, 59)

        # 将日期和时间合并为完整的datetime对象
        combined_datetime = datetime.combine(current_date, datetime.min.time()) + \
                            timedelta(hours=random_hour, minutes=random_minute, seconds=random_second)

        # 将日期和时间格式化为字符串
        formatted_date = current_date.strftime(date_format)
        formatted_time = combined_datetime.strftime(time_format)

        # 写入日期和时间数据
        sheet.cell(row=row_index, column=1, value=formatted_date)
        sheet.cell(row=row_index, column=2, value=formatted_time)

        row_index += 1

    # 增加一天，进行下一个日期的处理
    current_date += timedelta(days=1)

# 保存Excel文件
workbook.save('dates.xlsx')
